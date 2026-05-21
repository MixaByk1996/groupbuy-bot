"""Task 1: Catalog normalization.

Reads INPUT_products from the source workbook, normalizes records by
sku_norm, and writes the consolidated MASTER and ISSUES sheets back to a
filled workbook.

Run:
    python task1_normalize.py \
        --template templates/01_Task1_Catalog_Normalization.xls \
        --output 01_Task1_Catalog_Normalization.xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"
ET.register_namespace("", SS_NS)


def _cells(row: ET.Element) -> list[str]:
    """Return cell text values for a row, honoring ss:Index gaps."""
    values: list[str] = []
    expected = 1
    for cell in row.findall(f"{{{SS_NS}}}Cell"):
        idx = cell.get(f"{{{SS_NS}}}Index")
        if idx is not None:
            target = int(idx)
            while expected < target:
                values.append("")
                expected += 1
        data = cell.find(f"{{{SS_NS}}}Data")
        values.append("" if data is None or data.text is None else data.text)
        expected += 1
    return values


def _sheet_rows(workbook_xml: Path, sheet_name: str) -> list[list[str]]:
    tree = ET.parse(workbook_xml)
    root = tree.getroot()
    for sheet in root.findall(f"{{{SS_NS}}}Worksheet"):
        if sheet.get(f"{{{SS_NS}}}Name") != sheet_name:
            continue
        table = sheet.find(f"{{{SS_NS}}}Table")
        if table is None:
            return []
        return [_cells(row) for row in table.findall(f"{{{SS_NS}}}Row")]
    raise KeyError(f"Sheet not found: {sheet_name}")


# --- normalization rules ---

_WS_RE = re.compile(r"\s+")


def sku_norm(value: str) -> str:
    """UPPER + remove hyphens and whitespace."""
    if value is None:
        return ""
    return re.sub(r"[\s\-]+", "", value).upper()


def name_norm(value: str) -> str:
    """trim + collapse double spaces."""
    if value is None:
        return ""
    return _WS_RE.sub(" ", value).strip()


def color_norm(value: str, color_map: dict[str, str]) -> str:
    """Map color via COLOR_MAP, fall back to UPPER on match miss."""
    if not value:
        return ""
    raw = value.strip()
    if raw in color_map:
        return color_map[raw]
    lowered = raw.lower()
    for key, mapped in color_map.items():
        if key.lower() == lowered:
            return mapped
    return ""  # unknown color -> empty, will be flagged as MISSING_ATTR


def pick_photo(urls: list[str]) -> str:
    for url in urls:
        if url and url.strip():
            return url.strip()
    return ""


# --- consolidation ---


def consolidate(rows: list[dict], color_map: dict[str, str]):
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        key = sku_norm(row["sku"])
        groups[key].append(row)

    master: list[dict] = []
    issues: list[dict] = []

    for key, entries in sorted(groups.items()):
        names = {name_norm(e["name"]) for e in entries if e["name"]}
        name_value = sorted(names)[0] if names else ""

        colors = []
        for e in entries:
            c = color_norm(e["color"], color_map)
            if c:
                colors.append(c)
        color_value = colors[0] if colors else ""

        photo_value = pick_photo([e["photo_url"] for e in entries])

        flag = "OK"
        if len(entries) > 1:
            flag = "DUPLICATE"
        missing_attrs = []
        if not color_value:
            missing_attrs.append("color")
        if not photo_value:
            missing_attrs.append("photo_url")
        if missing_attrs and flag == "OK":
            flag = "MISSING_ATTR"

        master.append({
            "sku_norm": key,
            "name_norm": name_value,
            "color_norm": color_value,
            "photo_url_best": photo_value,
            "issues_flag": flag,
        })

        source_ids = ",".join(str(e["id"]) for e in entries)
        if len(entries) > 1:
            issues.append({
                "issue_type": "DUPLICATE",
                "sku_norm": key,
                "source_row_ids": source_ids,
                "details": f"Found {len(entries)} rows for the same normalized SKU.",
            })
        if missing_attrs:
            issues.append({
                "issue_type": "MISSING_ATTR",
                "sku_norm": key,
                "source_row_ids": source_ids,
                "details": "Missing: " + ", ".join(missing_attrs),
            })

    return master, issues


# --- workbook writing ---


def _set_header(ws, headers):
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDEEFF")
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(input_rows, master, issues, color_map, output_path: Path):
    wb = Workbook()
    # INPUT_products
    ws_in = wb.active
    ws_in.title = "INPUT_products"
    _set_header(ws_in, ["id", "sku", "name", "color", "photo_url"])
    for r in input_rows:
        ws_in.append([r["id"], r["sku"], r["name"], r["color"], r["photo_url"]])

    # MASTER
    ws_m = wb.create_sheet("MASTER")
    _set_header(ws_m, ["sku_norm", "name_norm", "color_norm", "photo_url_best", "issues_flag"])
    for row in master:
        ws_m.append([
            row["sku_norm"],
            row["name_norm"],
            row["color_norm"],
            row["photo_url_best"],
            row["issues_flag"],
        ])

    # ISSUES
    ws_i = wb.create_sheet("ISSUES")
    _set_header(ws_i, ["issue_type", "sku_norm", "source_row_ids", "details"])
    for row in issues:
        ws_i.append([
            row["issue_type"],
            row["sku_norm"],
            row["source_row_ids"],
            row["details"],
        ])

    # COLOR_MAP (preserved for transparency)
    ws_c = wb.create_sheet("COLOR_MAP")
    _set_header(ws_c, ["source_value", "color_norm"])
    for source, mapped in color_map.items():
        ws_c.append([source, mapped])

    # RULES
    ws_r = wb.create_sheet("RULES")
    _set_header(ws_r, ["Правило", "Описание"])
    for rule, desc in [
        ("sku_norm", "UPPER + убрать дефисы/пробелы"),
        ("name_norm", "trim + убрать двойные пробелы"),
        ("color_norm", "по COLOR_MAP (иначе пусто и MISSING_ATTR)"),
        ("photo_url_best", "выбрать непустое из дублей"),
        ("issues_flag", "OK / DUPLICATE / MISSING_ATTR"),
    ]:
        ws_r.append([rule, desc])

    # Column widths
    for ws in (ws_in, ws_m, ws_i, ws_c, ws_r):
        for col_idx, col_cells in enumerate(ws.columns, 1):
            longest = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                40, max(12, longest + 2)
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def load_input(template_path: Path):
    raw_rows = _sheet_rows(template_path, "INPUT_products")
    if not raw_rows:
        raise ValueError("INPUT_products sheet is empty")
    header = raw_rows[0]
    input_rows = []
    for r in raw_rows[1:]:
        record = dict(zip(header, r + [""] * (len(header) - len(r))))
        input_rows.append({
            "id": record.get("id", ""),
            "sku": record.get("sku", ""),
            "name": record.get("name", ""),
            "color": record.get("color", ""),
            "photo_url": record.get("photo_url", ""),
        })

    color_rows = _sheet_rows(template_path, "COLOR_MAP")
    color_map = {}
    for r in color_rows[1:]:
        if len(r) >= 2 and r[0]:
            color_map[r[0]] = r[1]
    return input_rows, color_map


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--template",
        type=Path,
        default=Path(__file__).parent / "templates" / "01_Task1_Catalog_Normalization.xls",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parent / "01_Task1_Catalog_Normalization.xlsx",
    )
    args = parser.parse_args()

    input_rows, color_map = load_input(args.template)
    master, issues = consolidate(input_rows, color_map)
    build_workbook(input_rows, master, issues, color_map, args.output)

    print(f"Wrote {args.output}")
    print(f"  MASTER rows: {len(master)}")
    print(f"  ISSUES rows: {len(issues)}")


if __name__ == "__main__":
    sys.exit(main() or 0)
