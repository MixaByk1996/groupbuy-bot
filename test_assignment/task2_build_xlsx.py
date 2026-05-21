"""Build 02_Task2_SQL_Checks_Template.xlsx with the 4 SQL queries.

The ANSWERS sheet content is sourced from `task2_sql_checks.sql` so the
authoritative answers live in a single, lint-friendly file.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HERE = Path(__file__).resolve().parent


def extract_queries(sql_path: Path) -> list[str]:
    text = sql_path.read_text(encoding="utf-8")
    # Split on the "-- #N — " query headers (keeping the marker).
    pattern = re.compile(r"(^|\n)-- #(\d+) —", re.MULTILINE)
    matches = list(pattern.finditer(text))
    queries: list[str] = []
    for idx, m in enumerate(matches):
        start = m.start() + (1 if m.group(1) == "\n" else 0)
        end = matches[idx + 1].start() + 1 if idx + 1 < len(matches) else len(text)
        block = text[start:end].rstrip()
        queries.append(block)
    assert len(queries) == 4, f"Expected 4 queries, got {len(queries)}"
    return queries


def build_workbook(queries: list[str], output: Path) -> None:
    wb = Workbook()

    # SCHEMA — copied from the template, for context.
    ws_s = wb.active
    ws_s.title = "SCHEMA"
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    for col, name in enumerate(["Таблица", "Поля"], 1):
        c = ws_s.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
    schema_rows = [
        ("orders",       "id, order_no, total_amount, currency, status"),
        ("order_lines",  "id, order_id, qty, price"),
        ("payments",     "id, order_id, amount, currency, status (paid/failed)"),
    ]
    for row in schema_rows:
        ws_s.append(row)
    ws_s.column_dimensions["A"].width = 18
    ws_s.column_dimensions["B"].width = 60

    # TASKS — from template.
    ws_t = wb.create_sheet("TASKS")
    for col, name in enumerate(["#", "Проверка", "Что вывести минимум"], 1):
        c = ws_t.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
    tasks = [
        (1, "Заказы без строк", "order_no, order_id, status"),
        (2, "Расхождение суммы заказа vs сумма строк (delta > 1)", "order_no, total_amount, lines_amount, delta"),
        (3, "Заказы paid, но paid_amount < total_amount", "order_no, total_amount, paid_amount, unpaid_amount"),
        (4, "Валютное несоответствие (paid)", "order_no, order_currency, payment_currency, amount"),
    ]
    for row in tasks:
        ws_t.append(row)
    ws_t.column_dimensions["A"].width = 5
    ws_t.column_dimensions["B"].width = 50
    ws_t.column_dimensions["C"].width = 50

    # ANSWERS — the deliverable.
    ws_a = wb.create_sheet("ANSWERS")
    for col, name in enumerate(["Блок", "SQL (PostgreSQL)"], 1):
        c = ws_a.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
    for i, sql in enumerate(queries, 1):
        ws_a.cell(row=i + 1, column=1, value=f"#{i} SQL").font = header_font
        cell = ws_a.cell(row=i + 1, column=2, value=sql)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_a.row_dimensions[i + 1].height = 220
    ws_a.column_dimensions["A"].width = 10
    ws_a.column_dimensions["B"].width = 110

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    queries = extract_queries(HERE / "task2_sql_checks.sql")
    output = HERE / "02_Task2_SQL_Checks_Template.xlsx"
    build_workbook(queries, output)
    print(f"Wrote {output} ({len(queries)} queries)")


if __name__ == "__main__":
    main()
