"""Task 3 — transform the ORDER_INPUT sheet and emit a print-ready invoice.

Pipeline:
    1. Read the ORDER_INPUT sheet from the source workbook.
    2. Compute ``line_total = qty * price * (1 - discount_pct / 100)`` and the
       order total.
    3. Fill the OUTPUT sheet with three blocks: partner, sale_order,
       sale_order_lines.
    4. Render ``invoice.html`` containing the header (no/date/customer/INN),
       the line items, and the total.

Run:
    python task3_transform_and_print.py \
        --template templates/03_Task3_Order_Print_Template.xls \
        --xlsx 03_Task3_Order_Print_Template.xlsx \
        --html invoice.html
"""
from __future__ import annotations

import argparse
import html
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"
HERE = Path(__file__).resolve().parent


# --- domain types ---


@dataclass
class OrderHeader:
    order_no: str
    date: str
    customer_name: str
    customer_inn: str
    currency: str


@dataclass
class OrderLine:
    sku: str
    name: str
    qty: Decimal
    price: Decimal
    discount_pct: Decimal

    @property
    def line_total(self) -> Decimal:
        gross = self.qty * self.price
        net = gross * (Decimal(1) - self.discount_pct / Decimal(100))
        return net.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


@dataclass
class Order:
    header: OrderHeader
    lines: list[OrderLine] = field(default_factory=list)

    @property
    def total(self) -> Decimal:
        return sum((l.line_total for l in self.lines), Decimal(0)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )


# --- reading the spreadsheet XML ---


def _cells(row: ET.Element) -> list[str]:
    values: list[str] = []
    expected = 1
    for cell in row.findall(f"{{{SS_NS}}}Cell"):
        idx = cell.get(f"{{{SS_NS}}}Index")
        if idx is not None:
            target = int(idx)
            while expected < target:
                values.append("")
                expected += 1
        data = cell.find(f"{{{SS_NS}}}Data")
        values.append("" if data is None or data.text is None else data.text)
        expected += 1
    return values


def _sheet_rows(workbook_xml: Path, sheet_name: str) -> list[list[str]]:
    tree = ET.parse(workbook_xml)
    root = tree.getroot()
    for sheet in root.findall(f"{{{SS_NS}}}Worksheet"):
        if sheet.get(f"{{{SS_NS}}}Name") != sheet_name:
            continue
        table = sheet.find(f"{{{SS_NS}}}Table")
        if table is None:
            return []
        return [_cells(row) for row in table.findall(f"{{{SS_NS}}}Row")]
    raise KeyError(f"Sheet not found: {sheet_name}")


def parse_order_input(template_path: Path) -> Order:
    """Parse the ORDER_INPUT sheet's two header/data blocks."""
    rows = _sheet_rows(template_path, "ORDER_INPUT")
    # The template has two header rows separated by a blank row.
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    for r in rows:
        if not any(c.strip() for c in r):
            if current:
                blocks.append(current)
                current = []
            continue
        current.append(r)
    if current:
        blocks.append(current)

    if len(blocks) < 2:
        raise ValueError("ORDER_INPUT must contain two blocks (header + lines)")

    head_block, line_block = blocks[0], blocks[1]
    head_keys, head_vals = head_block[0], head_block[1]
    head = dict(zip(head_keys, head_vals))
    header = OrderHeader(
        order_no=head.get("order_no", "").strip(),
        date=head.get("date", "").strip(),
        customer_name=head.get("customer_name", "").strip(),
        customer_inn=head.get("customer_inn", "").strip(),
        currency=head.get("currency", "").strip(),
    )

    line_keys = line_block[0]
    lines: list[OrderLine] = []
    for vals in line_block[1:]:
        record = dict(zip(line_keys, vals + [""] * (len(line_keys) - len(vals))))
        lines.append(OrderLine(
            sku=record.get("sku", "").strip(),
            name=record.get("name", "").strip(),
            qty=Decimal(record.get("qty", "0") or "0"),
            price=Decimal(record.get("price", "0") or "0"),
            discount_pct=Decimal(record.get("discount_pct", "0") or "0"),
        ))
    return Order(header=header, lines=lines)


# --- output workbook (OUTPUT sheet) ---


def _style_header(cells: list) -> None:
    for c in cells:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEEFF")
        c.alignment = Alignment(vertical="top", wrap_text=True)


def write_output_workbook(order: Order, xlsx_path: Path) -> None:
    wb = Workbook()

    # ORDER_INPUT — mirror the input for traceability.
    ws_in = wb.active
    ws_in.title = "ORDER_INPUT"
    _style_header([ws_in.cell(row=1, column=i + 1, value=v) for i, v in enumerate(
        ["order_no", "date", "customer_name", "customer_inn", "currency"]
    )])
    ws_in.append([
        order.header.order_no,
        order.header.date,
        order.header.customer_name,
        order.header.customer_inn,
        order.header.currency,
    ])
    ws_in.append([])
    _style_header([ws_in.cell(row=4, column=i + 1, value=v) for i, v in enumerate(
        ["sku", "name", "qty", "price", "discount_pct"]
    )])
    for l in order.lines:
        ws_in.append([l.sku, l.name, float(l.qty), float(l.price), float(l.discount_pct)])

    # OUTPUT
    ws_o = wb.create_sheet("OUTPUT")
    # 1) partner
    _style_header([ws_o.cell(row=1, column=i + 1, value=v) for i, v in enumerate(
        ["partner_inn", "partner_name"]
    )])
    ws_o.append([order.header.customer_inn, order.header.customer_name])
    ws_o.append([])

    # 2) sale_order
    _style_header([ws_o.cell(row=4, column=i + 1, value=v) for i, v in enumerate(
        ["sale_order_no", "sale_order_date", "currency", "partner_inn"]
    )])
    ws_o.append([
        order.header.order_no,
        order.header.date,
        order.header.currency,
        order.header.customer_inn,
    ])
    ws_o.append([])

    # 3) sale_order_lines
    _style_header([ws_o.cell(row=7, column=i + 1, value=v) for i, v in enumerate(
        ["order_no", "sku", "name", "qty", "price", "discount_pct", "line_total"]
    )])
    for l in order.lines:
        ws_o.append([
            order.header.order_no,
            l.sku,
            l.name,
            float(l.qty),
            float(l.price),
            float(l.discount_pct),
            float(l.line_total),
        ])
    # Total row.
    last_row = ws_o.max_row + 1
    ws_o.cell(row=last_row, column=6, value="TOTAL").font = Font(bold=True)
    total_cell = ws_o.cell(row=last_row, column=7, value=float(order.total))
    total_cell.font = Font(bold=True)

    # HINTS preserved.
    ws_h = wb.create_sheet("HINTS")
    _style_header([ws_h.cell(row=1, column=1, value="Подсказки")])
    for hint in [
        "Формула line_total: qty * price * (1 - discount_pct/100).",
        "Итог = сумма line_total по всем строкам.",
        "invoice.html должен содержать: номер/дата/контрагент+ИНН, строки, итог.",
    ]:
        ws_h.append([hint])

    # Column widths.
    for ws in (ws_in, ws_o, ws_h):
        for col_idx in range(1, ws.max_column + 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            longest = max(
                (len(str(c.value)) for c in ws[letter] if c.value is not None),
                default=12,
            )
            ws.column_dimensions[letter].width = min(36, max(12, longest + 2))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


# --- HTML invoice ---


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>Счёт {order_no}</title>
<style>
  body {{ font-family: "Segoe UI", Arial, sans-serif; margin: 32px; color: #222; }}
  h1 {{ margin: 0 0 8px; font-size: 22px; }}
  .meta {{ color: #555; margin-bottom: 24px; }}
  .meta div {{ margin: 2px 0; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; vertical-align: top; }}
  th {{ background: #f3f6fb; text-align: left; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  tfoot td {{ font-weight: bold; background: #fafafa; }}
  .partner {{ margin-bottom: 18px; }}
  .partner b {{ color: #111; }}
  @media print {{ body {{ margin: 12mm; }} }}
</style>
</head>
<body>
  <h1>Счёт {order_no}</h1>
  <div class="meta">
    <div>Дата: {date}</div>
    <div>Валюта: {currency}</div>
  </div>
  <div class="partner">
    <div><b>Контрагент:</b> {customer_name}</div>
    <div><b>ИНН:</b> {customer_inn}</div>
  </div>
  <table>
    <thead>
      <tr>
        <th>SKU</th>
        <th>Наименование</th>
        <th class="num">Кол-во</th>
        <th class="num">Цена</th>
        <th class="num">Скидка, %</th>
        <th class="num">Сумма</th>
      </tr>
    </thead>
    <tbody>
{rows}
    </tbody>
    <tfoot>
      <tr>
        <td colspan="5" class="num">Итого</td>
        <td class="num">{total} {currency}</td>
      </tr>
    </tfoot>
  </table>
</body>
</html>
"""


def _fmt(d: Decimal) -> str:
    return f"{d.normalize():f}" if d == d.to_integral_value() else f"{d:f}"


def render_invoice_html(order: Order) -> str:
    rows = []
    for l in order.lines:
        rows.append(
            "      <tr>"
            f"<td>{html.escape(l.sku)}</td>"
            f"<td>{html.escape(l.name)}</td>"
            f"<td class=\"num\">{_fmt(l.qty)}</td>"
            f"<td class=\"num\">{_fmt(l.price)}</td>"
            f"<td class=\"num\">{_fmt(l.discount_pct)}</td>"
            f"<td class=\"num\">{l.line_total:.2f}</td>"
            "</tr>"
        )
    return _HTML_TEMPLATE.format(
        order_no=html.escape(order.header.order_no),
        date=html.escape(order.header.date),
        currency=html.escape(order.header.currency),
        customer_name=html.escape(order.header.customer_name),
        customer_inn=html.escape(order.header.customer_inn),
        rows="\n".join(rows),
        total=f"{order.total:.2f}",
    )


# --- entry point ---


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--template",
        type=Path,
        default=HERE / "templates" / "03_Task3_Order_Print_Template.xls",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=HERE / "03_Task3_Order_Print_Template.xlsx",
    )
    parser.add_argument(
        "--html",
        type=Path,
        default=HERE / "invoice.html",
    )
    args = parser.parse_args(argv)

    order = parse_order_input(args.template)
    write_output_workbook(order, args.xlsx)
    args.html.write_text(render_invoice_html(order), encoding="utf-8")

    print(f"Wrote {args.xlsx}")
    print(f"Wrote {args.html}")
    print(f"  Lines: {len(order.lines)}; Total: {order.total} {order.header.currency}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
